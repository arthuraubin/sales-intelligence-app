import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill, NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime


def export_pitch_excel(df_products, df_stress=None, client_name="Client", filename="Pitch_Export.xlsx", custom_notes=""):
    """
    Exporte les données de produits, stress test et hypothèses dans un fichier Excel enrichi et bien formaté.

    Args:
        df_products (pd.DataFrame): Tableau des produits pricés
        df_stress (pd.DataFrame or None): Résultats détaillés du stress test (optionnel)
        client_name (str): Nom du client à inclure dans le nom du fichier
        filename (str): Nom du fichier Excel à générer (chemin possible)
        custom_notes (str): Commentaires personnalisés à inclure dans l’onglet Hypothèses
    """

    # Crée un dossier d'export si nécessaire
    export_folder = "exports"
    os.makedirs(export_folder, exist_ok=True) # crée un dossier, et tous les sous-dossiers nécessaires (s'ils n'existent pas)

    # Nettoyage du nom de fichier
    safe_name = client_name.replace(" ", "_").replace("/", "-")
    file_path = os.path.join(export_folder, filename or f"Pitch_{safe_name}.xlsx")

    wb = Workbook()

    # ---------------------------
    # Onglet 1 : Produits
    # ---------------------------
    ws_prod = wb.active
    ws_prod.title = "Produits"

    for r in dataframe_to_rows(df_products, index=False, header=True):
        ws_prod.append(r)

    # Mise en forme des colonnes
    for col in ws_prod.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

    # Ajustement automatique des colonnes
    for col in ws_prod.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws_prod.column_dimensions[col[0].column_letter].width = max_length + 2

    # Figer la ligne d'en-tête
    ws_prod.freeze_panes = "A2"

    # Table style
    tab = Table(displayName="Produits", ref=f"A1:{col[-1].column_letter}{ws_prod.max_row}")
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tab.tableStyleInfo = style
    ws_prod.add_table(tab)

    # ---------------------------
    # Onglet 2 : Stress Test
    # ---------------------------
    if df_stress is not None and not df_stress.empty:
        ws_stress = wb.create_sheet(title="Stress Test")
        for r in dataframe_to_rows(df_stress, index=False, header=True):
            ws_stress.append(r)

        # Mise en forme
        for col in ws_stress.iter_cols(min_row=1, max_row=1):
            for cell in col:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

        ws_stress.freeze_panes = "A2"

        for col in ws_stress.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws_stress.column_dimensions[col[0].column_letter].width = max_length + 2

    # ---------------------------
    # Onglet 3 : Hypothèses
    # ---------------------------
    ws_hypo = wb.create_sheet(title="Hypothèses")

    ws_hypo.append(["Date d'export", datetime.today().strftime("%d/%m/%Y")])
    ws_hypo.append(["Client", client_name])
    ws_hypo.append(["Modèle de pricing", "Black 76"])
    ws_hypo.append(["Courbes utilisées", "Zéro-coupon + volatilité implicite"])
    ws_hypo.append(["Hypothèses de stress test", "±50 bps sur les taux, ±20% sur la vol. "])

    if custom_notes:
        ws_hypo.append(["Notes personnalisées", custom_notes])

    for row in ws_hypo.iter_rows(min_row=1, max_row=ws_hypo.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    ws_hypo.column_dimensions['A'].width = 30
    ws_hypo.column_dimensions['B'].width = 70

    # ---------------------------
    # Finalisation
    # ---------------------------
    wb.save(file_path)
    print(f"✅ Export Excel terminé : {file_path}")
    return file_path